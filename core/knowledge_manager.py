"""
Lightweight local knowledge storage and retrieval for Billy.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _slugify(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return text or "folder"


@dataclass
class KnowledgeDocument:
    """Metadata for an uploaded knowledge document."""

    id: str
    folder_id: str
    filename: str
    stored_name: str
    extension: str
    mime_type: str
    uploaded_at: str
    chunk_count: int
    character_count: int


class KnowledgeManager:
    """Manage uploaded knowledge folders, metadata, and lightweight indexes."""

    SUPPORTED_EXTENSIONS = {
        ".pdf",
        ".txt",
        ".md",
        ".json",
        ".ini",
        ".csv",
        ".tsv",
        ".xlsx",
    }

    MIME_TYPES = {
        ".pdf": "application/pdf",
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".json": "application/json",
        ".ini": "text/plain",
        ".csv": "text/csv",
        ".tsv": "text/tab-separated-values",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    def __init__(self):
        self.root = Path("knowledge")
        self.manifest_path = self.root / "index.json"
        self.legacy_files_dir = self.root / "files"
        self.legacy_documents_path = self.root / "documents.json"
        self.legacy_index_path = self.root / "index.jsonl"
        self.root.mkdir(exist_ok=True)
        self._migrate_legacy_flat_layout_if_needed()
        self._ensure_default_folder()

    def list_folders(self, *, include_documents: bool = False) -> list[dict[str, Any]]:
        """Return folder metadata with summary counts."""
        folders = []
        for folder in self._load_manifest().get("folders", []):
            documents = self._load_folder_documents(folder)
            chunk_count = self._count_folder_chunks(folder)
            payload = {
                "id": folder["id"],
                "label": folder["label"],
                "path": folder["path"],
                "trigger_topics": folder.get("trigger_topics", []),
                "document_count": len(documents),
                "chunk_count": chunk_count,
                "created_at": folder.get("created_at"),
                "updated_at": folder.get("updated_at"),
            }
            if include_documents:
                payload["documents"] = sorted(
                    documents, key=lambda doc: doc["uploaded_at"], reverse=True
                )
            folders.append(payload)
        return sorted(folders, key=lambda item: item["label"].lower())

    def create_folder(
        self, *, label: str, trigger_topics: str | list[str] | None = None
    ) -> dict[str, Any]:
        """Create a new knowledge folder."""
        normalized_label = str(label or "").strip()
        if not normalized_label:
            raise ValueError("Folder name is required")

        manifest = self._load_manifest()
        existing_ids = {folder["id"] for folder in manifest.get("folders", [])}
        base_id = _slugify(normalized_label)
        folder_id = base_id
        suffix = 2
        while folder_id in existing_ids:
            folder_id = f"{base_id}-{suffix}"
            suffix += 1

        created_at = _utc_now_iso()
        folder = {
            "id": folder_id,
            "label": normalized_label,
            "path": folder_id,
            "trigger_topics": self._normalize_trigger_topics(trigger_topics),
            "meta_file": f"{folder_id}/meta.json",
            "documents_file": f"{folder_id}/documents.json",
            "index_file": f"{folder_id}/index.jsonl",
            "created_at": created_at,
            "updated_at": created_at,
        }
        manifest.setdefault("folders", []).append(folder)
        self._save_manifest(manifest)
        self._ensure_folder_layout(folder)
        self._save_folder_meta(folder)
        self._save_folder_documents(folder, [])
        return {
            "id": folder_id,
            "label": normalized_label,
            "trigger_topics": folder["trigger_topics"],
            "document_count": 0,
            "chunk_count": 0,
        }

    def update_folder(
        self,
        folder_id: str,
        *,
        label: str | None = None,
        trigger_topics: str | list[str] | None = None,
    ) -> dict[str, Any]:
        """Update folder label and trigger topics."""
        manifest = self._load_manifest()
        folder = self._get_manifest_folder(manifest, folder_id)
        if not folder:
            raise ValueError("Folder not found")

        if label is not None:
            normalized_label = str(label).strip()
            if not normalized_label:
                raise ValueError("Folder name is required")
            folder["label"] = normalized_label

        if trigger_topics is not None:
            folder["trigger_topics"] = self._normalize_trigger_topics(trigger_topics)

        folder["updated_at"] = _utc_now_iso()
        self._save_manifest(manifest)
        self._save_folder_meta(folder)
        documents = self._load_folder_documents(folder)
        return {
            "id": folder["id"],
            "label": folder["label"],
            "trigger_topics": folder.get("trigger_topics", []),
            "document_count": len(documents),
            "chunk_count": self._count_folder_chunks(folder),
        }

    def delete_folder(self, folder_id: str) -> bool:
        """Delete a knowledge folder and all files inside it."""
        manifest = self._load_manifest()
        folders = manifest.get("folders", [])
        folder = self._get_manifest_folder(manifest, folder_id)
        if not folder:
            return False

        if len(folders) <= 1:
            raise ValueError("At least one folder must remain")

        manifest["folders"] = [item for item in folders if item["id"] != folder_id]
        shutil.rmtree(self._folder_dir(folder), ignore_errors=True)
        self._save_manifest(manifest)
        return True

    def list_documents(self, folder_id: str | None = None) -> list[dict[str, Any]]:
        """Return known documents, optionally restricted to a folder."""
        folders = self._load_manifest().get("folders", [])
        documents: list[dict[str, Any]] = []
        for folder in folders:
            if folder_id and folder["id"] != folder_id:
                continue
            folder_docs = self._load_folder_documents(folder)
            for doc in folder_docs:
                doc["folder_label"] = folder["label"]
                doc["trigger_topics"] = folder.get("trigger_topics", [])
            documents.extend(folder_docs)
        return sorted(documents, key=lambda doc: doc["uploaded_at"], reverse=True)

    def add_document(
        self,
        *,
        folder_id: str,
        filename: str,
        data: bytes,
    ) -> dict[str, Any]:
        """Persist a knowledge file in the target folder and rebuild that folder index."""
        manifest = self._load_manifest()
        folder = self._get_manifest_folder(manifest, folder_id)
        if not folder:
            raise ValueError("Folder not found")

        extension = Path(filename).suffix.lower()
        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                "Unsupported file type. Supported types: "
                + ", ".join(sorted(self.SUPPORTED_EXTENSIONS))
            )

        self._ensure_folder_layout(folder)
        document_id = uuid.uuid4().hex
        stored_name = f"{document_id}{extension}"
        target_path = self._folder_files_dir(folder) / stored_name
        target_path.write_bytes(data)

        try:
            chunks, char_count = self._extract_chunks(target_path, extension)
        except Exception:
            target_path.unlink(missing_ok=True)
            raise

        document = KnowledgeDocument(
            id=document_id,
            folder_id=folder["id"],
            filename=filename,
            stored_name=stored_name,
            extension=extension,
            mime_type=self.MIME_TYPES.get(extension, "application/octet-stream"),
            uploaded_at=_utc_now_iso(),
            chunk_count=len(chunks),
            character_count=char_count,
        )

        documents = self._load_folder_documents(folder)
        documents.append(document.__dict__)
        self._save_folder_documents(folder, documents)
        folder["updated_at"] = _utc_now_iso()
        self._save_manifest(manifest)
        self._save_folder_meta(folder)
        self.rebuild_index(folder_id=folder["id"])

        payload = document.__dict__.copy()
        payload["folder_label"] = folder["label"]
        return payload

    def delete_document(self, document_id: str) -> bool:
        """Delete a knowledge document and refresh its folder index."""
        manifest = self._load_manifest()
        for folder in manifest.get("folders", []):
            documents = self._load_folder_documents(folder)
            remaining = []
            deleted = None
            for document in documents:
                if document["id"] == document_id:
                    deleted = document
                else:
                    remaining.append(document)

            if not deleted:
                continue

            (self._folder_files_dir(folder) / deleted["stored_name"]).unlink(
                missing_ok=True
            )
            self._save_folder_documents(folder, remaining)
            folder["updated_at"] = _utc_now_iso()
            self._save_manifest(manifest)
            self._save_folder_meta(folder)
            self.rebuild_index(folder_id=folder["id"])
            return True
        return False

    def rebuild_index(self, *, folder_id: str | None = None) -> dict[str, Any]:
        """Rebuild one folder index or all folder indexes."""
        manifest = self._load_manifest()
        folders = manifest.get("folders", [])
        target_folders = [
            folder
            for folder in folders
            if folder_id is None or folder["id"] == folder_id
        ]
        if folder_id and not target_folders:
            raise ValueError("Folder not found")

        indexed_documents = 0
        indexed_chunks = 0

        for folder in target_folders:
            self._ensure_folder_layout(folder)
            with self._folder_index_path(folder).open("w", encoding="utf-8") as handle:
                for document in self._load_folder_documents(folder):
                    path = self._folder_files_dir(folder) / document["stored_name"]
                    if not path.exists():
                        continue

                    chunks, _char_count = self._extract_chunks(
                        path, document["extension"]
                    )
                    indexed_documents += 1
                    for chunk in chunks:
                        payload = {
                            "id": f'{document["id"]}:{chunk["chunk_index"]}',
                            "folder_id": folder["id"],
                            "folder_label": folder["label"],
                            "document_id": document["id"],
                            "filename": document["filename"],
                            "chunk_index": chunk["chunk_index"],
                            "text": chunk["text"],
                        }
                        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
                        indexed_chunks += 1

        return {
            "ok": True,
            "folders_indexed": len(target_folders),
            "documents_indexed": indexed_documents,
            "chunks_indexed": indexed_chunks,
            "manifest_path": str(self.manifest_path),
        }

    def search(self, query: str, *, top_k: int = 3) -> dict[str, Any]:
        """Search knowledge indexes with folder trigger routing and fallback."""
        normalized_query = (query or "").strip()
        if not normalized_query:
            return {"ok": False, "error": "Query is required", "matches": []}

        tokens = self._tokenize(normalized_query)
        if not tokens:
            return {"ok": False, "error": "Query must contain words", "matches": []}

        folders = self._load_manifest().get("folders", [])
        if not folders:
            return {
                "ok": True,
                "query": normalized_query,
                "matches": [],
                "match_count": 0,
            }

        scored_folders = []
        for folder in folders:
            route_score = self._score_folder_route(normalized_query, tokens, folder)
            scored_folders.append((route_score, folder))

        candidate_folders = [folder for score, folder in scored_folders if score > 0]
        matches = self._search_folders(
            candidate_folders or folders,
            normalized_query,
            tokens,
            top_k,
            route_scores={folder["id"]: score for score, folder in scored_folders},
        )

        fallback_used = False
        if not matches and candidate_folders and len(candidate_folders) < len(folders):
            fallback_used = True
            matches = self._search_folders(
                folders,
                normalized_query,
                tokens,
                top_k,
                route_scores={folder["id"]: score for score, folder in scored_folders},
            )

        return {
            "ok": True,
            "query": normalized_query,
            "matches": matches,
            "match_count": len(matches),
            "searched_folders": [
                folder["label"] for folder in (candidate_folders or folders)
            ],
            "fallback_used": fallback_used,
        }

    def get_summary(self) -> dict[str, Any]:
        """Return an aggregate knowledge summary for UI display."""
        folders = self.list_folders(include_documents=False)
        return {
            "folder_count": len(folders),
            "document_count": sum(folder["document_count"] for folder in folders),
            "chunk_count": sum(folder["chunk_count"] for folder in folders),
        }

    def _search_folders(
        self,
        folders: list[dict[str, Any]],
        query: str,
        query_tokens: list[str],
        top_k: int,
        *,
        route_scores: dict[str, int],
    ) -> list[dict[str, Any]]:
        matches = []
        for folder in folders:
            folder_bonus = route_scores.get(folder["id"], 0)
            for row in self._load_index_rows(self._folder_index_path(folder)):
                score = self._score_chunk(
                    query=query,
                    query_tokens=query_tokens,
                    filename=str(row.get("filename") or ""),
                    text=str(row.get("text") or ""),
                )
                if score <= 0:
                    continue
                matches.append({
                    "folder_id": folder["id"],
                    "folder_label": folder["label"],
                    "document_id": row["document_id"],
                    "filename": row["filename"],
                    "chunk_index": row["chunk_index"],
                    "score": score + (folder_bonus * 0.35),
                    "snippet": self._build_snippet(
                        str(row.get("text") or ""), query_tokens
                    ),
                })

        matches.sort(
            key=lambda item: (
                -item["score"],
                item["folder_label"].lower(),
                item["filename"],
                item["chunk_index"],
            )
        )
        return matches[: max(1, min(top_k, 5))]

    def _score_folder_route(
        self, query: str, query_tokens: list[str], folder: dict[str, Any]
    ) -> int:
        score = 0
        folder_text = f'{folder["id"]} {folder["label"]}'.lower()
        query_lower = query.lower()

        if folder_text and folder_text in query_lower:
            score += 5

        for token in query_tokens:
            if token in folder_text:
                score += 1

        for topic in folder.get("trigger_topics", []):
            topic_text = str(topic).strip().lower()
            if not topic_text:
                continue
            if topic_text in query_lower:
                score += 6
                continue
            for token in self._tokenize(topic_text):
                if token in query_tokens:
                    score += 2
        return score

    def _extract_chunks(
        self, path: Path, extension: str
    ) -> tuple[list[dict[str, Any]], int]:
        text = self._extract_text(path, extension)
        if not text.strip():
            raise ValueError("No readable text could be extracted from this file")
        normalized_text = re.sub(r"\n{3,}", "\n\n", text).strip()
        chunks = self._chunk_text(normalized_text)
        return chunks, len(normalized_text)

    def _extract_text(self, path: Path, extension: str) -> str:
        if extension in {".txt", ".md", ".ini"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if extension == ".json":
            raw = path.read_text(encoding="utf-8", errors="ignore")
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                return raw
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        if extension in {".csv", ".tsv"}:
            dialect = "excel-tab" if extension == ".tsv" else "excel"
            with path.open(
                "r", encoding="utf-8", errors="ignore", newline=""
            ) as handle:
                reader = csv.reader(handle, dialect=dialect)
                rows = [
                    "\t".join(cell.strip() for cell in row if cell.strip())
                    for row in reader
                ]
            return "\n".join(row for row in rows if row)
        if extension == ".pdf":
            try:
                from pypdf import PdfReader
            except ImportError as exc:
                raise RuntimeError("PDF support requires the 'pypdf' package") from exc

            reader = PdfReader(str(path))
            pages = []
            for index, page in enumerate(reader.pages, start=1):
                page_text = page.extract_text() or ""
                if page_text.strip():
                    pages.append(f"Page {index}\n{page_text.strip()}")
            return "\n\n".join(pages)
        if extension == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError(
                    "Spreadsheet support requires the 'openpyxl' package"
                ) from exc

            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
            sections = []
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    values = [
                        str(cell).strip() for cell in row if cell not in (None, "")
                    ]
                    if values:
                        rows.append("\t".join(values))
                if rows:
                    sections.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
            return "\n\n".join(sections)
        raise ValueError(f"Unsupported file type: {extension}")

    def _chunk_text(
        self, text: str, *, chunk_size: int = 900, overlap: int = 150
    ) -> list[dict[str, Any]]:
        paragraphs = [part.strip() for part in text.split("\n\n") if part.strip()]
        if not paragraphs:
            return []

        chunks: list[dict[str, Any]] = []
        current = ""
        for paragraph in paragraphs:
            candidate = f"{current}\n\n{paragraph}".strip() if current else paragraph
            if len(candidate) <= chunk_size:
                current = candidate
                continue

            if current:
                chunks.append({"chunk_index": len(chunks), "text": current})
                tail = current[-overlap:].strip()
                current = f"{tail}\n\n{paragraph}".strip() if tail else paragraph
            else:
                for piece in self._split_large_paragraph(
                    paragraph, chunk_size, overlap
                ):
                    chunks.append({"chunk_index": len(chunks), "text": piece})
                current = ""

        if current:
            chunks.append({"chunk_index": len(chunks), "text": current})
        return chunks

    def _split_large_paragraph(
        self, paragraph: str, chunk_size: int, overlap: int
    ) -> list[str]:
        pieces = []
        start = 0
        while start < len(paragraph):
            end = min(len(paragraph), start + chunk_size)
            if end < len(paragraph):
                boundary = paragraph.rfind(" ", start, end)
                if boundary > start + 100:
                    end = boundary
            piece = paragraph[start:end].strip()
            if piece:
                pieces.append(piece)
            if end >= len(paragraph):
                break
            start = max(end - overlap, 0)
        return pieces

    def _load_manifest(self) -> dict[str, Any]:
        if not self.manifest_path.exists():
            return {"version": 2, "folders": []}
        try:
            return json.loads(self.manifest_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {"version": 2, "folders": []}

    def _save_manifest(self, manifest: dict[str, Any]):
        manifest["version"] = 2
        self.manifest_path.write_text(
            json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    def _get_manifest_folder(
        self, manifest: dict[str, Any], folder_id: str
    ) -> dict[str, Any] | None:
        return next(
            (
                folder
                for folder in manifest.get("folders", [])
                if folder["id"] == folder_id
            ),
            None,
        )

    def _ensure_default_folder(self):
        manifest = self._load_manifest()
        if manifest.get("folders"):
            for folder in manifest["folders"]:
                self._ensure_folder_layout(folder)
                self._save_folder_meta(folder)
                if not self._folder_documents_path(folder).exists():
                    self._save_folder_documents(folder, [])
            return
        self.create_folder(label="Shared")

    def _ensure_folder_layout(self, folder: dict[str, Any]):
        folder_dir = self._folder_dir(folder)
        folder_dir.mkdir(parents=True, exist_ok=True)
        self._folder_files_dir(folder).mkdir(exist_ok=True)

    def _save_folder_meta(self, folder: dict[str, Any]):
        meta = {
            "id": folder["id"],
            "label": folder["label"],
            "path": folder["path"],
            "trigger_topics": folder.get("trigger_topics", []),
            "created_at": folder.get("created_at"),
            "updated_at": folder.get("updated_at"),
        }
        self._folder_meta_path(folder).write_text(
            json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    def _load_folder_documents(self, folder: dict[str, Any]) -> list[dict[str, Any]]:
        path = self._folder_documents_path(folder)
        if not path.exists():
            return []
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []

    def _save_folder_documents(
        self, folder: dict[str, Any], documents: list[dict[str, Any]]
    ):
        self._folder_documents_path(folder).write_text(
            json.dumps(documents, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    def _count_folder_chunks(self, folder: dict[str, Any]) -> int:
        path = self._folder_index_path(folder)
        if not path.exists():
            return 0
        with path.open("r", encoding="utf-8") as handle:
            return sum(1 for _ in handle)

    def _load_index_rows(self, path: Path) -> list[dict[str, Any]]:
        if not path.exists():
            return []
        rows = []
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
        return rows

    def _folder_dir(self, folder: dict[str, Any]) -> Path:
        return self.root / folder["path"]

    def _folder_files_dir(self, folder: dict[str, Any]) -> Path:
        return self._folder_dir(folder) / "files"

    def _folder_documents_path(self, folder: dict[str, Any]) -> Path:
        return self.root / folder["documents_file"]

    def _folder_index_path(self, folder: dict[str, Any]) -> Path:
        return self.root / folder["index_file"]

    def _folder_meta_path(self, folder: dict[str, Any]) -> Path:
        return self.root / folder["meta_file"]

    def _normalize_trigger_topics(
        self, raw_topics: str | list[str] | None
    ) -> list[str]:
        if raw_topics is None:
            return []
        if isinstance(raw_topics, str):
            source = [part.strip() for part in raw_topics.split(",")]
        else:
            source = [str(part).strip() for part in raw_topics]
        seen: set[str] = set()
        normalized = []
        for topic in source:
            if not topic:
                continue
            lowered = topic.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            normalized.append(topic)
        return normalized

    def _score_chunk(
        self, *, query: str, query_tokens: list[str], filename: str, text: str
    ) -> float:
        haystack = f"{filename}\n{text}".lower()
        score = 0.0
        for token in query_tokens:
            pattern = re.compile(rf"\b{re.escape(token)}\b")
            matches = len(pattern.findall(haystack))
            if matches:
                score += min(matches, 5) * 2.0
        if query.lower() in haystack:
            score += 6.0
        if all(token in haystack for token in query_tokens):
            score += len(query_tokens) * 1.5
        return score

    def _build_snippet(self, text: str, tokens: list[str], *, window: int = 280) -> str:
        lowered = text.lower()
        first_hit = -1
        for token in tokens:
            idx = lowered.find(token)
            if idx != -1 and (first_hit == -1 or idx < first_hit):
                first_hit = idx

        if first_hit == -1:
            return text[:window].strip()

        start = max(first_hit - 80, 0)
        end = min(start + window, len(text))
        snippet = text[start:end].strip()
        if start > 0:
            snippet = "..." + snippet
        if end < len(text):
            snippet += "..."
        return snippet

    def _tokenize(self, text: str) -> list[str]:
        return [
            token for token in re.findall(r"[a-z0-9]+", text.lower()) if len(token) > 1
        ]

    def _migrate_legacy_flat_layout_if_needed(self):
        if self.manifest_path.exists():
            return

        legacy_docs_exist = self.legacy_documents_path.exists()
        legacy_files_exist = self.legacy_files_dir.exists()
        if not legacy_docs_exist and not legacy_files_exist:
            return

        created_at = _utc_now_iso()
        folder = {
            "id": "shared",
            "label": "Shared",
            "path": "shared",
            "trigger_topics": [],
            "meta_file": "shared/meta.json",
            "documents_file": "shared/documents.json",
            "index_file": "shared/index.jsonl",
            "created_at": created_at,
            "updated_at": created_at,
        }
        manifest = {"version": 2, "folders": [folder]}
        self._save_manifest(manifest)
        self._ensure_folder_layout(folder)

        documents = []
        if legacy_docs_exist:
            try:
                documents = json.loads(
                    self.legacy_documents_path.read_text(encoding="utf-8")
                )
            except json.JSONDecodeError:
                documents = []

        for document in documents:
            document["folder_id"] = "shared"

        if legacy_files_exist:
            for path in self.legacy_files_dir.iterdir():
                target = self._folder_files_dir(folder) / path.name
                if not target.exists():
                    shutil.move(str(path), str(target))
            shutil.rmtree(self.legacy_files_dir, ignore_errors=True)

        self._save_folder_documents(folder, documents)
        self._save_folder_meta(folder)
        self.rebuild_index(folder_id="shared")

        if self.legacy_documents_path.exists():
            self.legacy_documents_path.unlink(missing_ok=True)
        if self.legacy_index_path.exists():
            self.legacy_index_path.unlink(missing_ok=True)


knowledge_manager = KnowledgeManager()
